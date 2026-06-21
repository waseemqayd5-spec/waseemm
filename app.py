"""
نظام إدارة وكالة البشائر للأدوية والمستلزمات الطبية (جملة)
نسخة متطورة مع ميزات الذكاء الاصطناعي:
- التسعير الديناميكي
- كشف الشذوذ والاحتيال
- مساعد صيدلاني ذكي متقدم (RAG + ChatGPT)
- التعرف على الأدوية من الصور (OCR + Computer Vision + GPT-4 Vision)
- نظام توصيات مخصص للعملاء
- تحليل المشاعر من التقييمات
- استخراج بيانات الفواتير الورقية (OCR + NLP)
- لوحات معلومات ذكية (تحليل بيانات + استعلامات طبيعية)
"""

from flask import (
    Flask, request, jsonify, render_template_string, session,
    redirect, url_for, make_response, send_from_directory, Response, send_file
)
import os
import datetime
import hashlib
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import time
import json
import requests
import io
import re
import base64
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import threading
import queue
import random
import math
from collections import defaultdict
from datetime import timedelta
from dotenv import load_dotenv

# تحميل متغيرات البيئة
load_dotenv()

# =============================== تهيئة OpenAI ===============================
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("⚠️ openai غير مثبت، قم بتثبيته: pip install openai")

OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')
OPENAI_MODEL = os.environ.get('OPENAI_MODEL', 'gpt-3.5-turbo')
if OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY
    print("✅ OpenAI API تم تهيئته بنجاح")
else:
    print("⚠️ OPENAI_API_KEY غير موجود في متغيرات البيئة")

# محاولة استيراد مكتبات الذكاء الاصطناعي (مع fallback)
try:
    import numpy as np
    from sklearn.ensemble import IsolationForest
    from sklearn.preprocessing import StandardScaler
    from sklearn.metrics.pairwise import cosine_similarity
    from sklearn.feature_extraction.text import TfidfVectorizer
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    print("⚠️ scikit-learn غير مثبت، بعض ميزات الذكاء الاصطناعي ستكون محدودة")
    print("   قم بتثبيته: pip install scikit-learn numpy")

try:
    import pytesseract
    from PIL import Image
    import cv2
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("⚠️ pytesseract أو PIL غير مثبت، ميزة التعرف من الصور غير متوفرة")
    print("   قم بتثبيته: pip install pytesseract pillow opencv-python")

try:
    from sentence_transformers import SentenceTransformer
    SENTENCE_TRANSFORMER_AVAILABLE = True
except ImportError:
    SENTENCE_TRANSFORMER_AVAILABLE = False
    print("⚠️ sentence-transformers غير مثبت، ميزة RAG محدودة")
    print("   قم بتثبيته: pip install sentence-transformers")

# =============================== التهيئة ===============================
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
UPLOAD_FOLDER = 'static/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

DATABASE_URL = os.environ.get('DATABASE_URL', None)
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')

# =============================== دوال قاعدة البيانات ===============================
def get_db_connection():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    else:
        os.makedirs('data', exist_ok=True)
        conn = sqlite3.connect('data/supermarket.db')
        conn.row_factory = sqlite3.Row
    return conn

def execute_query(query, params=None, fetch_one=False, fetch_all=False, commit=False):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if params is None:
            params = ()
        cur.execute(query, params)
        if commit:
            conn.commit()
        if fetch_one:
            return cur.fetchone()
        elif fetch_all:
            return cur.fetchall()
    finally:
        cur.close()
        conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# =============================== نظام الصلاحيات ===============================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session or session.get('role') not in allowed_roles:
                return "غير مصرح (صلاحية غير كافية)", 403
            return f(*args, **kwargs)
        return decorated
    return decorator

admin_required = role_required(['admin'])
pharmacist_required = role_required(['admin', 'pharmacist'])
cashier_required = role_required(['admin', 'cashier'])
store_keeper_required = role_required(['admin', 'store_keeper'])
purchaser_required = role_required(['admin', 'purchaser'])

# =============================== دوال مساعدة ===============================
def get_settings():
    rows = execute_query("SELECT key, value FROM settings", fetch_all=True)
    settings = {row['key']: row['value'] for row in rows} if rows else {}
    defaults = {
        'company_name': 'وكالة البشائر للأدوية والمستلزمات الطبية',
        'company_logo': '💊',
        'company_address': 'اليمن - صنعاء',
        'company_phone': '967771602370',
        'company_whatsapp': '967771602370',
        'delivery_fee': '0.00',
        'points_per_riyal': '0',
        'points_redeem_rate': '0',
        'currency': 'ريال يمني',
        'tax_rate': '0.00',
        'default_unit': 'حبة'
    }
    defaults.update(settings)
    return defaults

def save_settings(settings_dict):
    for key, value in settings_dict.items():
        if DATABASE_URL:
            execute_query("INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                          (key, value), commit=True)
        else:
            execute_query("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                          (key, value), commit=True)

def log_inventory(product_id, product_name, change_type, quantity_change, old_qty, new_qty, notes, user_id):
    execute_query("""
        INSERT INTO inventory_logs (product_id, product_name, change_type, quantity_change, old_quantity, new_quantity, notes, user_id, timestamp)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (product_id, product_name, change_type, quantity_change, old_qty, new_qty, notes, user_id, datetime.datetime.now().isoformat()), commit=True)

# =============================== إنشاء الجداول ===============================
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    if DATABASE_URL:
        # PostgreSQL
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash VARCHAR(200) NOT NULL,
                role VARCHAR(20) DEFAULT 'cashier',
                full_name VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key VARCHAR(100) PRIMARY KEY,
                value TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id SERIAL PRIMARY KEY,
                phone VARCHAR(20) UNIQUE,
                name VARCHAR(100),
                address TEXT,
                tax_number VARCHAR(50),
                commercial_register VARCHAR(50),
                payment_terms VARCHAR(50) DEFAULT 'cash',
                loyalty_points INTEGER DEFAULT 0,
                total_spent REAL DEFAULT 0,
                visits INTEGER DEFAULT 0,
                last_visit DATE,
                tier VARCHAR(20) DEFAULT 'عادي',
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                phone VARCHAR(20),
                address TEXT,
                email VARCHAR(100),
                contact_person VARCHAR(100),
                license_number VARCHAR(50),
                bank_account TEXT,
                notes TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS units (
                id SERIAL PRIMARY KEY,
                name VARCHAR(50) UNIQUE NOT NULL,
                symbol VARCHAR(10),
                base_unit_id INTEGER REFERENCES units(id),
                conversion_factor REAL DEFAULT 1.0,
                is_base BOOLEAN DEFAULT FALSE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                barcode VARCHAR(50) UNIQUE,
                name VARCHAR(200) NOT NULL,
                description TEXT,
                category VARCHAR(50),
                unit_id INTEGER REFERENCES units(id),
                price REAL NOT NULL,
                cost_price REAL,
                quantity REAL DEFAULT 0,
                min_quantity REAL DEFAULT 10,
                supplier_id INTEGER REFERENCES suppliers(id),
                expiry_date DATE,
                added_date DATE,
                last_updated DATE,
                image_url TEXT,
                image_url2 TEXT,
                is_active INTEGER DEFAULT 1,
                active_ingredient TEXT,
                dosage_form TEXT,
                strength TEXT,
                manufacturer TEXT,
                batch_number TEXT,
                storage_conditions TEXT,
                prescription_required BOOLEAN DEFAULT FALSE,
                unit_pack_size INTEGER,
                dynamic_price REAL,
                price_updated_at TIMESTAMP,
                sales_velocity REAL DEFAULT 0,
                abc_class CHAR(1) DEFAULT 'C',
                xyz_class CHAR(1) DEFAULT 'Z'
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS purchases (
                id SERIAL PRIMARY KEY,
                supplier_id INTEGER REFERENCES suppliers(id),
                invoice_number VARCHAR(50),
                total_cost REAL,
                purchase_date DATE,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS purchase_items (
                id SERIAL PRIMARY KEY,
                purchase_id INTEGER REFERENCES purchases(id),
                product_id INTEGER REFERENCES products(id),
                quantity REAL,
                cost_price REAL,
                total REAL,
                batch_number TEXT,
                expiry_date DATE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS offers (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                description TEXT,
                code VARCHAR(50),
                discount_type VARCHAR(20) DEFAULT 'percentage',
                discount_value REAL,
                start_date DATE,
                end_date DATE,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                invoice_number VARCHAR(50) UNIQUE NOT NULL,
                customer_id INTEGER REFERENCES customers(id),
                customer_name VARCHAR(100),
                customer_phone VARCHAR(20),
                customer_address TEXT,
                total REAL NOT NULL,
                discount REAL DEFAULT 0,
                final_total REAL NOT NULL,
                payment_method VARCHAR(50),
                points_earned INTEGER DEFAULT 0,
                points_used INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES users(id),
                is_anomaly BOOLEAN DEFAULT FALSE,
                anomaly_score REAL DEFAULT 0,
                anomaly_reason TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoice_items (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER REFERENCES invoices(id),
                product_id INTEGER REFERENCES products(id),
                product_name VARCHAR(200),
                quantity REAL NOT NULL,
                price REAL NOT NULL,
                total REAL NOT NULL,
                batch_number TEXT,
                expiry_date DATE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventory_logs (
                id SERIAL PRIMARY KEY,
                product_id INTEGER REFERENCES products(id),
                product_name TEXT,
                change_type TEXT,
                quantity_change REAL,
                old_quantity REAL,
                new_quantity REAL,
                notes TEXT,
                user_id INTEGER REFERENCES users(id),
                timestamp TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS returns (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER REFERENCES invoices(id),
                product_id INTEGER REFERENCES products(id),
                product_name VARCHAR(200),
                quantity REAL,
                price REAL,
                total REAL,
                reason TEXT,
                return_date DATE,
                created_by INTEGER REFERENCES users(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS alerts (
                id SERIAL PRIMARY KEY,
                type VARCHAR(50),
                product_id INTEGER REFERENCES products(id),
                message TEXT,
                is_read BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER REFERENCES customers(id),
                customer_name VARCHAR(100),
                customer_phone VARCHAR(20),
                rating INTEGER CHECK (rating >= 1 AND rating <= 5),
                comment TEXT,
                sentiment_score REAL,
                sentiment_label VARCHAR(20),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS price_history (
                id SERIAL PRIMARY KEY,
                product_id INTEGER REFERENCES products(id),
                old_price REAL,
                new_price REAL,
                reason TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS anomaly_logs (
                id SERIAL PRIMARY KEY,
                invoice_id INTEGER REFERENCES invoices(id),
                anomaly_score REAL,
                reason TEXT,
                is_reviewed BOOLEAN DEFAULT FALSE,
                reviewed_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS recommendations (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER REFERENCES customers(id),
                product_id INTEGER REFERENCES products(id),
                score REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS product_similarity (
                product_id1 INTEGER REFERENCES products(id),
                product_id2 INTEGER REFERENCES products(id),
                similarity_score REAL,
                PRIMARY KEY (product_id1, product_id2)
            )
        """)
    else:
        # SQLite
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'cashier',
                full_name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phone TEXT UNIQUE,
                name TEXT,
                address TEXT,
                tax_number TEXT,
                commercial_register TEXT,
                payment_terms TEXT DEFAULT 'cash',
                loyalty_points INTEGER DEFAULT 0,
                total_spent REAL DEFAULT 0,
                visits INTEGER DEFAULT 0,
                last_visit DATE,
                tier TEXT DEFAULT 'عادي',
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                email TEXT,
                contact_person TEXT,
                license_number TEXT,
                bank_account TEXT,
                notes TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS units (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                symbol TEXT,
                base_unit_id INTEGER REFERENCES units(id),
                conversion_factor REAL DEFAULT 1.0,
                is_base BOOLEAN DEFAULT FALSE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                barcode TEXT UNIQUE,
                name TEXT NOT NULL,
                description TEXT,
                category TEXT,
                unit_id INTEGER REFERENCES units(id),
                price REAL NOT NULL,
                cost_price REAL,
                quantity REAL DEFAULT 0,
                min_quantity REAL DEFAULT 10,
                supplier_id INTEGER REFERENCES suppliers(id),
                expiry_date DATE,
                added_date DATE,
                last_updated DATE,
                image_url TEXT,
                image_url2 TEXT,
                is_active INTEGER DEFAULT 1,
                active_ingredient TEXT,
                dosage_form TEXT,
                strength TEXT,
                manufacturer TEXT,
                batch_number TEXT,
                storage_conditions TEXT,
                prescription_required BOOLEAN DEFAULT FALSE,
                unit_pack_size INTEGER,
                dynamic_price REAL,
                price_updated_at TIMESTAMP,
                sales_velocity REAL DEFAULT 0,
                abc_class CHAR(1) DEFAULT 'C',
                xyz_class CHAR(1) DEFAULT 'Z'
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS purchases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id INTEGER REFERENCES suppliers(id),
                invoice_number TEXT,
                total_cost REAL,
                purchase_date DATE,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS purchase_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                purchase_id INTEGER REFERENCES purchases(id),
                product_id INTEGER REFERENCES products(id),
                quantity REAL,
                cost_price REAL,
                total REAL,
                batch_number TEXT,
                expiry_date DATE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS offers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                code TEXT,
                discount_type TEXT DEFAULT 'percentage',
                discount_value REAL,
                start_date DATE,
                end_date DATE,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_number TEXT UNIQUE NOT NULL,
                customer_id INTEGER REFERENCES customers(id),
                customer_name TEXT,
                customer_phone TEXT,
                customer_address TEXT,
                total REAL NOT NULL,
                discount REAL DEFAULT 0,
                final_total REAL NOT NULL,
                payment_method TEXT,
                points_earned INTEGER DEFAULT 0,
                points_used INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER REFERENCES users(id),
                is_anomaly BOOLEAN DEFAULT FALSE,
                anomaly_score REAL DEFAULT 0,
                anomaly_reason TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS invoice_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER REFERENCES invoices(id),
                product_id INTEGER REFERENCES products(id),
                product_name TEXT,
                quantity REAL NOT NULL,
                price REAL NOT NULL,
                total REAL NOT NULL,
                batch_number TEXT,
                expiry_date DATE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventory_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER REFERENCES products(id),
                product_name TEXT,
                change_type TEXT,
                quantity_change REAL,
                old_quantity REAL,
                new_quantity REAL,
                notes TEXT,
                user_id INTEGER REFERENCES users(id),
                timestamp TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER REFERENCES invoices(id),
                product_id INTEGER REFERENCES products(id),
                product_name TEXT,
                quantity REAL,
                price REAL,
                total REAL,
                reason TEXT,
                return_date DATE,
                created_by INTEGER REFERENCES users(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT,
                product_id INTEGER REFERENCES products(id),
                message TEXT,
                is_read BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER REFERENCES customers(id),
                customer_name TEXT,
                customer_phone TEXT,
                rating INTEGER,
                comment TEXT,
                sentiment_score REAL,
                sentiment_label TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS price_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER REFERENCES products(id),
                old_price REAL,
                new_price REAL,
                reason TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS anomaly_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER REFERENCES invoices(id),
                anomaly_score REAL,
                reason TEXT,
                is_reviewed BOOLEAN DEFAULT FALSE,
                reviewed_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS recommendations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER REFERENCES customers(id),
                product_id INTEGER REFERENCES products(id),
                score REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS product_similarity (
                product_id1 INTEGER REFERENCES products(id),
                product_id2 INTEGER REFERENCES products(id),
                similarity_score REAL,
                PRIMARY KEY (product_id1, product_id2)
            )
        """)

    # إدخال البيانات الافتراضية
    cur.execute("SELECT COUNT(*) FROM units")
    if cur.fetchone()[0] == 0:
        units = [
            ('حبة', 'حب', None, 1.0, 1),
            ('علبة', 'علبة', None, 1.0, 0),
            ('شريط', 'شريط', None, 1.0, 0),
            ('كرتونة', 'كرتونة', None, 1.0, 0),
            ('لتر', 'لتر', None, 1.0, 0),
            ('ملجم', 'ملجم', None, 1.0, 0),
            ('جرام', 'جرام', None, 1.0, 0),
        ]
        for unit in units:
            cur.execute("INSERT INTO units (name, symbol, base_unit_id, conversion_factor, is_base) VALUES (?, ?, ?, ?, ?)", unit)

    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        hashed = generate_password_hash('admin123')
        cur.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                    ('admin', hashed, 'admin', 'مدير النظام'))
        hashed_ph = generate_password_hash('pharma123')
        cur.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                    ('pharmacist', hashed_ph, 'pharmacist', 'صيدلي رئيسي'))
        hashed_cashier = generate_password_hash('cashier123')
        cur.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                    ('cashier', hashed_cashier, 'cashier', 'كاشير'))
        hashed_stock = generate_password_hash('stock123')
        cur.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                    ('stock', hashed_stock, 'store_keeper', 'أمين مخزن'))
        hashed_purch = generate_password_hash('purch123')
        cur.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                    ('purchaser', hashed_purch, 'purchaser', 'مندوب مشتريات'))

    cur.execute("SELECT COUNT(*) FROM settings")
    if cur.fetchone()[0] == 0:
        default_settings = [
            ('company_name', 'وكالة البشائر للأدوية والمستلزمات الطبية'),
            ('company_logo', '💊'),
            ('company_address', 'اليمن - صنعاء'),
            ('company_phone', '967771602370'),
            ('company_whatsapp', '967771602370'),
            ('delivery_fee', '0.00'),
            ('points_per_riyal', '0'),
            ('points_redeem_rate', '0'),
            ('currency', 'ريال يمني'),
            ('tax_rate', '0.00'),
            ('default_unit', 'حبة')
        ]
        for key, val in default_settings:
            cur.execute("INSERT INTO settings (key, value) VALUES (?, ?)", (key, val))

    cur.execute("SELECT COUNT(*) FROM suppliers")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO suppliers (name, phone, address, contact_person, license_number) VALUES (?, ?, ?, ?, ?)",
                    ('شركة الحكمة للأدوية', '0500000001', 'صنعاء', 'أحمد القرشي', 'LIC001'))
        cur.execute("INSERT INTO suppliers (name, phone, address, contact_person, license_number) VALUES (?, ?, ?, ?, ?)",
                    ('مستودع الأمان', '0500000002', 'عدن', 'سامي النجار', 'LIC002'))

    cur.execute("SELECT COUNT(*) FROM customers")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO customers (phone, name, address, tax_number, commercial_register, payment_terms) VALUES (?, ?, ?, ?, ?, ?)",
                    ('0500000000', 'صيدلية الرحمة', 'صنعاء - شارع الستين', 'TAX001', 'CR001', 'credit'))
        cur.execute("INSERT INTO customers (phone, name, address, tax_number, commercial_register, payment_terms) VALUES (?, ?, ?, ?, ?, ?)",
                    ('0500000003', 'مستشفى السلام', 'صنعاء - السبعين', 'TAX002', 'CR002', 'cash'))

    cur.execute("SELECT COUNT(*) FROM products")
    if cur.fetchone()[0] == 0:
        today = datetime.date.today()
        cur.execute("SELECT id FROM units WHERE name='حبة'")
        unit_id = cur.fetchone()[0]
        cur.execute("SELECT id FROM suppliers WHERE name LIKE '%الحكمة%'")
        sup1 = cur.fetchone()[0]
        cur.execute("SELECT id FROM suppliers WHERE name LIKE '%الأمان%'")
        sup2 = cur.fetchone()[0]

        products = [
            ("6281234567890", "باراسيتامول 500 مجم", "مسكن وخافض حرارة يستخدم لعلاج الآلام الخفيفة والمتوسطة والحمى.", "مسكنات", unit_id, 5.0, 3.5, 200, 20, sup1, (today + datetime.timedelta(days=365)).isoformat(), today.isoformat(), today.isoformat(), None, None, 1, "باراسيتامول", "أقراص", "500 مجم", "شركة الحكمة", "BATCH001", "درجة حرارة الغرفة", False, 10, None, None, 0, 'A', 'X'),
            ("6280987654321", "أموكسيسيلين 500 مجم", "مضاد حيوي واسع الطيف لعلاج الالتهابات البكتيرية.", "مضادات حيوية", unit_id, 15.0, 10.0, 100, 15, sup1, (today + datetime.timedelta(days=180)).isoformat(), today.isoformat(), today.isoformat(), None, None, 1, "أموكسيسيلين", "كبسولات", "500 مجم", "جلوبال فارما", "BATCH002", "يحفظ في الثلاجة", True, 20, None, None, 0, 'A', 'X'),
            ("6281122334455", "أوميبرازول 40 مجم", "علاج قرحة المعدة والارتجاع المريئي.", "جهاز هضمي", unit_id, 8.0, 5.0, 150, 10, sup2, (today + datetime.timedelta(days=200)).isoformat(), today.isoformat(), today.isoformat(), None, None, 1, "أوميبرازول", "أقراص", "40 مجم", "الأمان", "BATCH003", "درجة حرارة الغرفة", False, 14, None, None, 0, 'B', 'Y'),
            ("6289988776655", "سيتالوبريم 20 مجم", "مضاد اكتئاب من مجموعة مثبطات استرداد السيروتونين.", "أمراض نفسية", unit_id, 12.0, 8.0, 80, 10, sup2, (today + datetime.timedelta(days=150)).isoformat(), today.isoformat(), today.isoformat(), None, None, 1, "سيتالوبريم", "أقراص", "20 مجم", "الأمان", "BATCH004", "درجة حرارة الغرفة", True, 28, None, None, 0, 'B', 'Z'),
        ]
        for prod in products:
            cur.execute("""
                INSERT INTO products (barcode, name, description, category, unit_id, price, cost_price, quantity, min_quantity, supplier_id, expiry_date, added_date, last_updated, image_url, image_url2, is_active, active_ingredient, dosage_form, strength, manufacturer, batch_number, storage_conditions, prescription_required, unit_pack_size, dynamic_price, price_updated_at, sales_velocity, abc_class, xyz_class)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, prod)

    conn.commit()
    conn.close()

init_db()

# =============================== نظام الإشعارات (SSE) ===============================
notification_queue = queue.Queue()

def check_alerts():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, quantity, min_quantity FROM products
        WHERE quantity <= min_quantity AND is_active = 1
    """)
    low_stock = cur.fetchall()
    for row in low_stock:
        msg = f"المنتج '{row['name']}' مخزونه منخفض (الكمية: {row['quantity']}، الحد الأدنى: {row['min_quantity']})"
        execute_query("INSERT INTO alerts (type, product_id, message) VALUES (?, ?, ?)",
                      ('low_stock', row['id'], msg), commit=True)
        notification_queue.put({'type': 'low_stock', 'message': msg, 'product_id': row['id']})

    today = datetime.date.today()
    expiry_threshold = today + datetime.timedelta(days=30)
    cur.execute("""
        SELECT id, name, expiry_date FROM products
        WHERE expiry_date <= ? AND expiry_date >= ? AND is_active = 1
    """, (expiry_threshold, today))
    near_expiry = cur.fetchall()
    for row in near_expiry:
        days_left = (row['expiry_date'] - today).days
        msg = f"المنتج '{row['name']}' سينتهي صلاحيته خلال {days_left} يوم (تاريخ الصلاحية: {row['expiry_date']})"
        execute_query("INSERT INTO alerts (type, product_id, message) VALUES (?, ?, ?)",
                      ('expiry', row['id'], msg), commit=True)
        notification_queue.put({'type': 'expiry', 'message': msg, 'product_id': row['id']})
    conn.close()

def alert_scheduler():
    while True:
        check_alerts()
        time.sleep(600)

threading.Thread(target=alert_scheduler, daemon=True).start()

@app.route('/api/events')
def sse_events():
    def event_stream():
        while True:
            try:
                data = notification_queue.get(timeout=30)
                yield f"data: {json.dumps(data)}\n\n"
            except queue.Empty:
                yield ": heartbeat\n\n"
    return Response(event_stream(), mimetype="text/event-stream")

# =============================== رفع الصور ===============================
@app.route('/api/upload-image', methods=['POST'])
@login_required
def upload_image():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'لا يوجد ملف'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'اسم الملف فارغ'})
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        timestamp = int(time.time())
        new_filename = f"{name}_{timestamp}{ext}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_filename))
        return jsonify({'success': True, 'url': f'/static/uploads/{new_filename}'})
    return jsonify({'success': False, 'message': 'نوع الملف غير مسموح'})

# =============================== ميزة 1: التسعير الديناميكي ===============================
def calculate_dynamic_price(product_id):
    """حساب السعر الديناميكي للمنتج بناءً على عوامل متعددة"""
    product = execute_query("""
        SELECT id, name, price, cost_price, quantity, sales_velocity, expiry_date, category, abc_class
        FROM products WHERE id = ?
    """, (product_id,), fetch_one=True)
    if not product:
        return None

    base_price = product['price']
    cost = product['cost_price'] or base_price * 0.6
    quantity = product['quantity'] or 0
    velocity = product['sales_velocity'] or 1.0
    expiry = product['expiry_date']

    factors = {
        'demand': 1.0,
        'stock': 1.0,
        'expiry': 1.0,
        'category': 1.0,
        'class': 1.0
    }

    if velocity > 5:
        factors['demand'] = 1.08
    elif velocity > 2:
        factors['demand'] = 1.03
    elif velocity < 0.5:
        factors['demand'] = 0.95

    if quantity < 20:
        factors['stock'] = 1.05
    elif quantity > 200:
        factors['stock'] = 0.92

    if expiry:
        today = datetime.date.today()
        if isinstance(expiry, str):
            expiry = datetime.datetime.strptime(expiry, '%Y-%m-%d').date()
        days_left = (expiry - today).days
        if days_left < 30:
            factors['expiry'] = 0.80
        elif days_left < 60:
            factors['expiry'] = 0.90

    if product['category'] in ['مسكنات', 'مضادات حيوية']:
        factors['category'] = 1.05
    elif product['category'] in ['أمراض مزمنة', 'أمراض نفسية']:
        factors['category'] = 1.02

    if product['abc_class'] == 'A':
        factors['class'] = 1.03
    elif product['abc_class'] == 'C':
        factors['class'] = 0.95

    dynamic_price = base_price
    for factor in factors.values():
        dynamic_price *= factor

    min_price = cost * 1.1
    if dynamic_price < min_price:
        dynamic_price = min_price

    dynamic_price = round(dynamic_price * 2) / 2

    execute_query("""
        UPDATE products SET dynamic_price = ?, price_updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (dynamic_price, product_id), commit=True)

    execute_query("""
        INSERT INTO price_history (product_id, old_price, new_price, reason, created_by)
        VALUES (?, ?, ?, ?, ?)
    """, (product_id, base_price, dynamic_price, 'تحديث تلقائي - تسعير ديناميكي', 1), commit=True)

    return dynamic_price

@app.route('/api/dynamic-price/<int:product_id>', methods=['GET'])
def api_dynamic_price(product_id):
    price = calculate_dynamic_price(product_id)
    if price is None:
        return jsonify({'success': False, 'message': 'المنتج غير موجود'})
    return jsonify({'success': True, 'price': price})

# =============================== ميزة 2: كشف الشذوذ والاحتيال ===============================
def detect_anomaly(invoice_data):
    """كشف الشذوذ في الفواتير باستخدام Isolation Forest"""
    if not SKLEARN_AVAILABLE:
        return {'is_anomaly': False, 'score': 0, 'reason': 'مكتبة scikit-learn غير مثبتة'}

    try:
        past_invoices = execute_query("""
            SELECT id, total, discount, final_total,
                   (SELECT COUNT(*) FROM invoice_items WHERE invoice_id = invoices.id) as item_count
            FROM invoices
            WHERE created_at >= datetime('now', '-30 days')
            ORDER BY created_at DESC LIMIT 100
        """, fetch_all=True)

        if not past_invoices or len(past_invoices) < 10:
            return {'is_anomaly': False, 'score': 0, 'reason': 'بيانات تدريب غير كافية'}

        X = []
        for inv in past_invoices:
            X.append([
                inv['total'] or 0,
                inv['discount'] or 0,
                inv['final_total'] or 0,
                inv['item_count'] or 0
            ])

        X.append([
            invoice_data.get('total', 0),
            invoice_data.get('discount', 0),
            invoice_data.get('final_total', 0),
            invoice_data.get('item_count', 0)
        ])

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        model = IsolationForest(contamination=0.1, random_state=42)
        predictions = model.fit_predict(X_scaled)
        scores = model.decision_function(X_scaled)

        is_anomaly = predictions[-1] == -1
        score = float(scores[-1])

        reason = "طبيعي"
        if is_anomaly:
            reasons = []
            if invoice_data.get('total', 0) > 10000:
                reasons.append("قيمة الفاتورة مرتفعة جداً")
            if invoice_data.get('discount', 0) / max(invoice_data.get('total', 1), 1) > 0.5:
                reasons.append("نسبة الخصم مرتفعة جداً")
            if invoice_data.get('item_count', 0) > 30:
                reasons.append("عدد الأصناف كبير جداً")
            if not reasons:
                reasons.append("نمط غير عادي")
            reason = "، ".join(reasons)

        return {'is_anomaly': is_anomaly, 'score': score, 'reason': reason}

    except Exception as e:
        return {'is_anomaly': False, 'score': 0, 'reason': f'خطأ في التحليل: {str(e)}'}

@app.route('/api/check-anomaly', methods=['POST'])
def check_anomaly():
    data = request.json
    cart = data.get('cart', [])
    total = data.get('total', 0)
    discount = data.get('discount', 0)
    final_total = data.get('final_total', total - discount)

    invoice_data = {
        'total': total,
        'discount': discount,
        'final_total': final_total,
        'item_count': len(cart)
    }

    result = detect_anomaly(invoice_data)
    return jsonify({
        'success': True,
        'is_anomaly': result['is_anomaly'],
        'score': result['score'],
        'reason': result['reason']
    })

# =============================== ميزة 3: المساعد الذكي المتقدم (RAG + ChatGPT) ===============================
# تحميل نموذج التضمينات إذا كان متوفراً
embedding_model = None
if SENTENCE_TRANSFORMER_AVAILABLE:
    try:
        embedding_model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
    except:
        embedding_model = None

def get_product_embeddings():
    """توليد تضمينات لجميع المنتجات للاستخدام في RAG"""
    products = execute_query("""
        SELECT id, name, description, active_ingredient, category, manufacturer,
               strength, dosage_form, price
        FROM products WHERE is_active = 1
    """, fetch_all=True)

    if not products or embedding_model is None:
        return None, None

    texts = []
    for p in products:
        text = f"{p['name']} {p['description'] or ''} {p['active_ingredient'] or ''} {p['category'] or ''} {p['manufacturer'] or ''}"
        texts.append(text)

    try:
        embeddings = embedding_model.encode(texts)
        return products, embeddings
    except:
        return None, None

def rag_search(query, top_k=3):
    """البحث باستخدام RAG (استرجاع معزز بالتوليد)"""
    products, embeddings = get_product_embeddings()
    if products is None or embeddings is None:
        return []

    try:
        query_embedding = embedding_model.encode([query])
        similarities = cosine_similarity(query_embedding, embeddings)[0]
        top_indices = similarities.argsort()[-top_k:][::-1]

        results = []
        for idx in top_indices:
            results.append({
                'product': dict(products[idx]),
                'similarity': float(similarities[idx])
            })
        return results
    except:
        return []

@app.route('/api/chat', methods=['POST'])
def api_chat():
    """المساعد الذكي المتقدم مع RAG و ChatGPT"""
    data = request.json
    user_message = data.get('message', '').strip()
    if not user_message:
        return jsonify({"success": False, "error": "رسالة فارغة"})

    # 1. البحث باستخدام RAG
    rag_results = rag_search(user_message, top_k=5)

    # 2. بناء السياق من النتائج
    context = ""
    if rag_results:
        context = "المنتجات ذات الصلة:\n"
        for r in rag_results:
            p = r['product']
            context += f"- {p['name']} (المادة الفعالة: {p['active_ingredient'] or 'غير محدد'}, السعر: {p['price']} ريال, التشابه: {r['similarity']:.2f})\n"

    # 3. استخدام ChatGPT (OpenAI)
    if OPENAI_AVAILABLE and OPENAI_API_KEY:
        try:
            system_prompt = f"""أنت مساعد صيدلاني محترف في وكالة البشائر للأدوية والمستلزمات الطبية.
معلومات إضافية من قاعدة البيانات:
{context}

تعليمات مهمة:
- أجب باللغة العربية الفصحى
- قدم معلومات دقيقة عن الأدوية
- لا تقدم استشارات طبية تشخيصية
- إذا سأل عن دواء معين، استخدم المعلومات المتوفرة أعلاه
- إذا لم تعرف الإجابة، قل ذلك بصراحة
- كن ودوداً ومحترفاً"""

            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=500,
                temperature=0.7
            )
            reply = response.choices[0].message.content
            return jsonify({"success": True, "reply": reply, "rag_results": rag_results, "model": "ChatGPT"})
        except Exception as e:
            print(f"OpenAI error: {e}")

    # 4. استخدام Gemini كبديل (إذا كان المفتاح موجوداً)
    if GEMINI_API_KEY:
        try:
            system_prompt = f"""أنت مساعد صيدلاني في وكالة البشائر للأدوية والمستلزمات الطبية.
            معلومات إضافية من قاعدة البيانات:
            {context}

            تعليمات مهمة:
            - أجب باللغة العربية الفصحى
            - قدم معلومات دقيقة عن الأدوية
            - لا تقدم استشارات طبية تشخيصية
            - إذا سأل عن دواء معين، استخدم المعلومات المتوفرة أعلاه
            - إذا لم تعرف الإجابة، قل ذلك بصراحة

            سؤال العميل: {user_message}"""

            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"
            payload = {
                "contents": [{"role": "user", "parts": [{"text": system_prompt}]}]
            }
            resp = requests.post(url, json=payload, timeout=15)
            if resp.status_code == 200:
                candidates = resp.json().get("candidates", [])
                if candidates and candidates[0].get("content", {}).get("parts"):
                    reply = candidates[0]["content"]["parts"][0]["text"]
                    return jsonify({"success": True, "reply": reply, "rag_results": rag_results, "model": "Gemini"})
        except Exception as e:
            print(f"Gemini error: {e}")

    # 5. الرد المحلي المحسن (في حال فشل جميع الخيارات)
    msg_lower = user_message.lower()

    for r in rag_results[:3]:
        p = r['product']
        if p['name'].lower() in msg_lower:
            reply = f"🔍 **{p['name']}**\n"
            reply += f"💊 المادة الفعالة: {p['active_ingredient'] or 'غير محدد'}\n"
            reply += f"📦 الفئة: {p['category'] or 'غير محدد'}\n"
            reply += f"🏭 الشركة المصنعة: {p['manufacturer'] or 'غير محدد'}\n"
            reply += f"💰 السعر: {p['price']} ريال\n"
            reply += f"📝 الوصف: {p['description'] or 'لا يوجد وصف'}\n"
            if p['strength']:
                reply += f"⚡ التركيز: {p['strength']}\n"
            if p['dosage_form']:
                reply += f"💊 الشكل الصيدلاني: {p['dosage_form']}\n"
            return jsonify({"success": True, "reply": reply, "rag_results": rag_results, "model": "local"})

    if rag_results:
        reply = "📋 **منتجات ذات صلة بسؤالك:**\n"
        for i, r in enumerate(rag_results[:3], 1):
            p = r['product']
            reply += f"{i}. {p['name']} - {p['price']} ريال"
            if p['active_ingredient']:
                reply += f" (المادة الفعالة: {p['active_ingredient']})"
            reply += "\n"
        reply += "\n💡 هل تريد معرفة المزيد عن أي من هذه المنتجات؟"
    else:
        reply = "🔍 لم أجد منتجات تطابق سؤالك بالضبط. يمكنك البحث عن دواء معين أو سؤال عن فئة دوائية محددة."

    return jsonify({"success": True, "reply": reply, "rag_results": rag_results, "model": "local"})

# =============================== ميزة 4: تحليل الصور باستخدام GPT-4 Vision ===============================
@app.route('/api/analyze-medicine-image', methods=['POST'])
@login_required
def analyze_medicine_image():
    """تحليل صورة دواء باستخدام GPT-4 Vision"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'لا يوجد ملف'})
    
    file = request.files['file']
    if not OPENAI_AVAILABLE or not OPENAI_API_KEY:
        return jsonify({'success': False, 'message': 'مفتاح OpenAI غير موجود'})
    
    try:
        # قراءة الصورة وتحويلها إلى base64
        image_data = base64.b64encode(file.read()).decode('utf-8')
        
        response = openai.ChatCompletion.create(
            model="gpt-4o",  # أو gpt-4-vision-preview
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "حلل هذه الصورة لدواء أو عبوة دواء. استخرج: اسم الدواء، المادة الفعالة، التركيز، الشركة المصنعة، أي معلومات أخرى مهمة. أجب باللغة العربية."},
                        {"type": "image_url", "image_url": f"data:image/jpeg;base64,{image_data}"}
                    ]
                }
            ],
            max_tokens=500
        )
        
        analysis = response.choices[0].message.content
        return jsonify({'success': True, 'analysis': analysis})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# =============================== ميزة 5: التعرف على الأدوية من الصور (OCR) ===============================
@app.route('/api/scan-image', methods=['POST'])
@login_required
def scan_image():
    """تحليل صورة الدواء واستخراج المعلومات"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'لا يوجد ملف'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'اسم الملف فارغ'})

    if not OCR_AVAILABLE:
        return jsonify({'success': False, 'message': 'مكتبات OCR غير مثبتة'})

    try:
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', filename)
        file.save(temp_path)

        image = cv2.imread(temp_path)
        if image is None:
            return jsonify({'success': False, 'message': 'لا يمكن قراءة الصورة'})

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        text = pytesseract.image_to_string(gray, lang='ara+eng')
        text = text.strip()

        if not text:
            return jsonify({'success': False, 'message': 'لم يتم العثور على نص في الصورة'})

        barcode = None
        try:
            from pyzbar import pyzbar
            barcodes = pyzbar.decode(image)
            if barcodes:
                barcode = barcodes[0].data.decode('utf-8')
        except:
            pass

        products = []
        if barcode:
            product = execute_query("SELECT * FROM products WHERE barcode = ? AND is_active = 1", (barcode,), fetch_one=True)
            if product:
                products.append(dict(product))

        if not products:
            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if len(line) > 3:
                    results = execute_query("""
                        SELECT * FROM products
                        WHERE name LIKE ? AND is_active = 1
                        LIMIT 5
                    """, (f"%{line}%",), fetch_all=True)
                    for r in results:
                        if r not in products:
                            products.append(dict(r))

        try:
            os.remove(temp_path)
        except:
            pass

        return jsonify({
            'success': True,
            'text': text,
            'barcode': barcode,
            'products': products[:5],
            'message': f'تم العثور على {len(products)} منتج (منتجات)'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'خطأ: {str(e)}'})

# =============================== ميزة 6: نظام توصيات مخصص للعملاء ===============================
def generate_recommendations(customer_id):
    """توليد توصيات مخصصة للعميل"""
    customer_items = execute_query("""
        SELECT DISTINCT ii.product_id
        FROM invoice_items ii
        JOIN invoices i ON ii.invoice_id = i.id
        WHERE i.customer_id = ?
    """, (customer_id,), fetch_all=True)

    if not customer_items:
        top_products = execute_query("""
            SELECT product_id, SUM(quantity) as total_sold
            FROM invoice_items
            GROUP BY product_id
            ORDER BY total_sold DESC
            LIMIT 5
        """, fetch_all=True)
        return [p['product_id'] for p in top_products]

    bought_ids = [p['product_id'] for p in customer_items]

    similar_products = execute_query("""
        SELECT product_id2 as product_id, similarity_score
        FROM product_similarity
        WHERE product_id1 IN ({})
        ORDER BY similarity_score DESC
        LIMIT 10
    """.format(','.join('?' * len(bought_ids))), bought_ids, fetch_all=True)

    if similar_products:
        recommended = [p['product_id'] for p in similar_products if p['product_id'] not in bought_ids]
        if recommended:
            return recommended[:5]

    categories = execute_query("""
        SELECT DISTINCT category FROM products WHERE id IN ({})
    """.format(','.join('?' * len(bought_ids))), bought_ids, fetch_all=True)

    if categories:
        cats = [c['category'] for c in categories if c['category']]
        if cats:
            placeholders = ','.join('?' * len(cats))
            cat_products = execute_query(f"""
                SELECT id FROM products
                WHERE category IN ({placeholders}) AND is_active = 1
                AND id NOT IN ({','.join('?' * len(bought_ids))})
                ORDER BY sales_velocity DESC, quantity DESC
                LIMIT 5
            """, cats + bought_ids, fetch_all=True)
            return [p['id'] for p in cat_products]

    top = execute_query("""
        SELECT product_id, SUM(quantity) as total_sold
        FROM invoice_items
        GROUP BY product_id
        ORDER BY total_sold DESC
        LIMIT 5
    """, fetch_all=True)
    return [p['product_id'] for p in top]

@app.route('/api/recommendations/<int:customer_id>')
def api_recommendations(customer_id):
    recs = generate_recommendations(customer_id)

    if recs:
        placeholders = ','.join('?' * len(recs))
        products = execute_query(f"""
            SELECT id, name, price, image_url, category, active_ingredient
            FROM products WHERE id IN ({placeholders}) AND is_active = 1
        """, recs, fetch_all=True)
    else:
        products = []

    for p_id in recs[:5]:
        execute_query("""
            INSERT OR REPLACE INTO recommendations (customer_id, product_id, score, created_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        """, (customer_id, p_id, 1.0), commit=True)

    return jsonify({
        'success': True,
        'recommendations': [dict(p) for p in products]
    })

@app.route('/api/calculate-similarity', methods=['POST'])
@admin_required
def calculate_similarity():
    if not SKLEARN_AVAILABLE:
        return jsonify({'success': False, 'message': 'مكتبة scikit-learn غير مثبتة'})

    products = execute_query("""
        SELECT id, name, description, category, active_ingredient, manufacturer
        FROM products WHERE is_active = 1
    """, fetch_all=True)

    if len(products) < 2:
        return jsonify({'success': False, 'message': 'تحتاج إلى منتجين على الأقل'})

    texts = []
    for p in products:
        text = f"{p['name']} {p['description'] or ''} {p['category'] or ''} {p['active_ingredient'] or ''} {p['manufacturer'] or ''}"
        texts.append(text)

    vectorizer = TfidfVectorizer(max_features=100, stop_words=None)
    tfidf = vectorizer.fit_transform(texts)
    similarity_matrix = cosine_similarity(tfidf)

    count = 0
    for i in range(len(products)):
        for j in range(i+1, len(products)):
            score = float(similarity_matrix[i][j])
            if score > 0.1:
                execute_query("""
                    INSERT OR REPLACE INTO product_similarity (product_id1, product_id2, similarity_score)
                    VALUES (?, ?, ?)
                """, (products[i]['id'], products[j]['id'], score), commit=True)
                count += 1

    return jsonify({
        'success': True,
        'message': f'تم حساب {count} علاقة تشابه بين {len(products)} منتج'
    })

# =============================== ميزة 7: تحليل المشاعر من التقييمات ===============================
def analyze_sentiment(text):
    """تحليل المشاعر باستخدام ChatGPT أو نموذج بسيط"""
    if not text:
        return {'label': 'محايد', 'score': 0.0}

    # استخدام ChatGPT إذا كان متاحاً
    if OPENAI_AVAILABLE and OPENAI_API_KEY:
        try:
            prompt = f"""حلل المشاعر في النص التالي وأعط النتيجة كـ JSON:
النص: "{text}"
أخرج JSON بالشكل التالي: {{"label": "إيجابي|سلبي|محايد", "score": 0.0-1.0}}"""

            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=100,
                temperature=0.3
            )
            content = response.choices[0].message.content
            json_match = re.search(r'\{.*\}', content)
            if json_match:
                result = json.loads(json_match.group())
                return {'label': result.get('label', 'محايد'), 'score': result.get('score', 0.5)}
        except:
            pass

    # تحليل بسيط باستخدام الكلمات المفتاحية (fallback)
    positive_words = ['ممتاز', 'رائع', 'جيد', 'مفيد', 'فعال', 'سريع', 'دقيق', 'قوي', 'مذهل', 'ممتازة', 'جيدة']
    negative_words = ['سيء', 'مضر', 'ضعيف', 'بطيء', 'غير مفيد', 'خطير', 'مكلف', 'رديء', 'سيئة', 'مشكلة', 'أخطاء']

    text_lower = text.lower()
    pos_score = sum(1 for w in positive_words if w in text_lower)
    neg_score = sum(1 for w in negative_words if w in text_lower)

    if pos_score > neg_score:
        label = 'إيجابي'
        score = min(0.5 + (pos_score - neg_score) * 0.1, 1.0)
    elif neg_score > pos_score:
        label = 'سلبي'
        score = min(0.5 + (neg_score - pos_score) * 0.1, 1.0)
    else:
        label = 'محايد'
        score = 0.5

    return {'label': label, 'score': score}

@app.route('/api/feedback', methods=['POST'])
def submit_feedback():
    data = request.json
    customer_id = data.get('customer_id')
    customer_name = data.get('customer_name', '')
    customer_phone = data.get('customer_phone', '')
    rating = data.get('rating', 3)
    comment = data.get('comment', '')

    if not comment and rating < 1:
        return jsonify({'success': False, 'message': 'يرجى كتابة تعليق أو تقييم'})

    sentiment = analyze_sentiment(comment)

    execute_query("""
        INSERT INTO feedback (customer_id, customer_name, customer_phone, rating, comment, sentiment_score, sentiment_label)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (customer_id, customer_name, customer_phone, rating, comment, sentiment['score'], sentiment['label']), commit=True)

    return jsonify({
        'success': True,
        'message': 'شكراً لتقييمك',
        'sentiment': sentiment
    })

# =============================== ميزة 8: استخراج بيانات الفواتير الورقية (OCR + NLP) ===============================
@app.route('/api/analyze-invoice', methods=['POST'])
@login_required
def analyze_invoice():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'لا يوجد ملف'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'اسم الملف فارغ'})

    if not OCR_AVAILABLE:
        return jsonify({'success': False, 'message': 'مكتبات OCR غير مثبتة'})

    try:
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', filename)
        file.save(temp_path)

        image = cv2.imread(temp_path)
        if image is None:
            return jsonify({'success': False, 'message': 'لا يمكن قراءة الصورة'})

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        text = pytesseract.image_to_string(gray, lang='ara+eng')
        text = text.strip()

        if not text:
            return jsonify({'success': False, 'message': 'لم يتم العثور على نص في الصورة'})

        extracted = {
            'invoice_number': None,
            'supplier_name': None,
            'date': None,
            'total': None,
            'items': []
        }

        invoice_match = re.search(r'(?:رقم|فاتورة|invoice|#)\s*[: ]?\s*([A-Za-z0-9\-]+)', text, re.IGNORECASE)
        if invoice_match:
            extracted['invoice_number'] = invoice_match.group(1)

        supplier_match = re.search(r'(?:مورد|شركة|supplier|from)\s*[: ]?\s*([^\n]+)', text, re.IGNORECASE)
        if supplier_match:
            extracted['supplier_name'] = supplier_match.group(1).strip()

        date_match = re.search(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', text)
        if not date_match:
            date_match = re.search(r'(\d{4}[/\-]\d{1,2}[/\-]\d{1,2})', text)
        if date_match:
            extracted['date'] = date_match.group(1)

        total_match = re.search(r'(?:إجمالي|total|المجموع|المبلغ)\s*[: ]?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
        if total_match:
            extracted['total'] = float(total_match.group(1).replace(',', ''))

        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            item_match = re.search(r'([^\d]+)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s*$', line)
            if item_match:
                extracted['items'].append({
                    'name': item_match.group(1).strip(),
                    'quantity': float(item_match.group(2)),
                    'price': float(item_match.group(3))
                })

        if not extracted['items']:
            products = execute_query("SELECT name FROM products WHERE is_active = 1", fetch_all=True)
            for p in products:
                if p['name'] in text:
                    extracted['items'].append({
                        'name': p['name'],
                        'quantity': 1,
                        'price': 0
                    })

        try:
            os.remove(temp_path)
        except:
            pass

        return jsonify({
            'success': True,
            'text': text[:500],
            'extracted': extracted,
            'message': 'تم استخراج البيانات بنجاح'
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'خطأ: {str(e)}'})

@app.route('/api/save-invoice-data', methods=['POST'])
@admin_required
def save_invoice_data():
    data = request.json
    try:
        supplier_name = data.get('supplier_name', '')
        supplier = None
        if supplier_name:
            supplier = execute_query("SELECT id FROM suppliers WHERE name LIKE ?", (f"%{supplier_name}%",), fetch_one=True)

        purchase_id = None
        if data.get('items'):
            total = data.get('total', 0)
            if total == 0:
                total = sum(item['quantity'] * item['price'] for item in data['items'])

            execute_query("""
                INSERT INTO purchases (supplier_id, invoice_number, total_cost, purchase_date, notes)
                VALUES (?, ?, ?, ?, ?)
            """, (
                supplier['id'] if supplier else None,
                data.get('invoice_number', ''),
                total,
                data.get('date') or datetime.date.today().isoformat(),
                'تم الاستيراد من الفاتورة الورقية'
            ), commit=True)

            purchase = execute_query("SELECT id FROM purchases ORDER BY id DESC LIMIT 1", fetch_one=True)
            purchase_id = purchase['id']

            for item in data['items']:
                product = execute_query("SELECT id FROM products WHERE name LIKE ?", (f"%{item['name']}%",), fetch_one=True)
                if product:
                    execute_query("""
                        INSERT INTO purchase_items (purchase_id, product_id, quantity, cost_price, total)
                        VALUES (?, ?, ?, ?, ?)
                    """, (purchase_id, product['id'], item['quantity'], item['price'], item['quantity'] * item['price']), commit=True)

                    execute_query("UPDATE products SET quantity = quantity + ? WHERE id = ?",
                                  (item['quantity'], product['id']), commit=True)

        return jsonify({'success': True, 'message': 'تم حفظ الفاتورة بنجاح', 'purchase_id': purchase_id})

    except Exception as e:
        return jsonify({'success': False, 'message': f'خطأ: {str(e)}'})

# =============================== ميزة 9: لوحات معلومات ذكية (تحليل بيانات + استعلامات طبيعية) ===============================
@app.route('/api/ai-insights', methods=['POST'])
@admin_required
def ai_insights():
    """الحصول على رؤى وتحليلات ذكية من البيانات باستخدام ChatGPT"""
    data = request.json
    query_type = data.get('type', 'general')
    
    stats = execute_query("""
        SELECT 
            (SELECT COUNT(*) FROM products WHERE is_active = 1) as total_products,
            (SELECT COUNT(*) FROM customers) as total_customers,
            (SELECT IFNULL(SUM(final_total), 0) FROM invoices) as total_revenue,
            (SELECT COUNT(*) FROM invoices) as total_invoices,
            (SELECT IFNULL(AVG(final_total), 0) FROM invoices) as avg_order_value
    """, fetch_one=True)
    
    if not OPENAI_AVAILABLE or not OPENAI_API_KEY:
        return jsonify({'success': False, 'message': 'مفتاح OpenAI غير موجود'})
    
    prompt = f"""
أنت محلل أعمال ذكي. إليك بيانات وكالة أدوية:
- عدد المنتجات: {stats['total_products']}
- عدد العملاء: {stats['total_customers']}
- إجمالي الإيرادات: {stats['total_revenue']} ريال
- عدد الفواتير: {stats['total_invoices']}
- متوسط قيمة الطلب: {stats['avg_order_value']} ريال

بناءً على هذه البيانات، قدم تحليلاً ذكياً وتوصيات استراتيجية لتحسين الأعمال.
استخدم اللغة العربية في الرد.
"""
    
    try:
        response = openai.ChatCompletion.create(
            model=OPENAI_MODEL,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500,
            temperature=0.7
        )
        return jsonify({'success': True, 'insights': response.choices[0].message.content})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/analytics-query', methods=['POST'])
@admin_required
def analytics_query():
    """معالجة الاستعلامات الطبيعية للتحليل"""
    data = request.json
    query = data.get('query', '').strip()

    if not query:
        return jsonify({'success': False, 'message': 'الاستعلام فارغ'})

    # محاولة استخدام ChatGPT لتحويل الاستعلام
    if OPENAI_AVAILABLE and OPENAI_API_KEY:
        try:
            prompt = f"""أنت مساعد تحليل بيانات. المستخدم يسأل عن بيانات المبيعات.
قم بتحليل السؤال التالي واستخراج المعلومات المطلوبة من قاعدة البيانات.
إذا كان السؤال عن أفضل منتج، أو إيرادات، أو مبيعات، قم بإنشاء استعلام SQL مناسب.
السؤال: "{query}"

أخرج استجابة منظمة كـ JSON مع:
- type: نوع الطلب (top_product, revenue, sales_count, general)
- sql: استعلام SQL إذا أمكن
- explanation: شرح بالعربية للنتيجة

فقط أخرج JSON ولا شيء آخر."""

            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=300,
                temperature=0.3
            )
            content = response.choices[0].message.content
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                try:
                    analysis = json.loads(json_match.group())
                    if analysis.get('sql'):
                        result = execute_query(analysis['sql'], fetch_all=True)
                        if result:
                            return jsonify({
                                'success': True,
                                'result': f"{analysis.get('explanation', 'النتيجة:')}\n\n" +
                                          '\n'.join([str(dict(r)) for r in result[:20]])
                            })
                except:
                    pass
        except:
            pass

    # رد محلي بسيط (fallback)
    query_lower = query.lower()
    result = ""

    if 'أفضل' in query_lower or 'top' in query_lower:
        if 'منتج' in query_lower or 'دواء' in query_lower:
            top = execute_query("""
                SELECT p.name, SUM(ii.quantity) as total_sold
                FROM invoice_items ii JOIN products p ON ii.product_id = p.id
                GROUP BY ii.product_id ORDER BY total_sold DESC LIMIT 5
            """, fetch_all=True)
            if top:
                result = "🏆 أفضل 5 منتجات مبيعاً:\n" + '\n'.join([f"- {p['name']}: {p['total_sold']} وحدة" for p in top])
    elif 'إيراد' in query_lower or 'revenue' in query_lower:
        revenue = execute_query("SELECT IFNULL(SUM(final_total), 0) as total FROM invoices", fetch_one=True)
        result = f"💰 إجمالي الإيرادات: {revenue['total']} ريال"
    elif 'مبيعات' in query_lower or 'sales' in query_lower:
        sales = execute_query("SELECT COUNT(*) as count FROM invoices", fetch_one=True)
        result = f"📊 عدد الفواتير: {sales['count']} فاتورة"

    if not result:
        result = "🔍 لم أتمكن من فهم السؤال. حاول أن تسأل عن: أفضل المنتجات، الإيرادات، عدد المبيعات، أو تصنيف الفئات."

    return jsonify({'success': True, 'result': result})

# =============================== ميزة 10: واجهة ChatGPT في لوحة التحكم ===============================
@app.route('/admin/ai-chat')
@admin_required
def ai_chat_page():
    """صفحة الدردشة مع ChatGPT مع سياق الأعمال"""
    return render_template_string("""
    <!DOCTYPE html>
    <html dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>المساعد الذكي المتقدم</title>
        <style>
            :root {
                --bg: #000;
                --text: #FFD700;
                --card-bg: #111;
                --border: #FFD700;
                --btn-bg: #FFD700;
                --btn-text: #000;
            }
            body.light {
                --bg: #f5f5f5;
                --text: #000;
                --card-bg: #fff;
                --border: #007bff;
                --btn-bg: #007bff;
                --btn-text: #fff;
            }
            body{background:var(--bg);color:var(--text);padding:20px;font-family:Arial;transition:background 0.3s,color 0.3s;}
            .container{max-width:800px;margin:auto;}
            .header{background:var(--card-bg);border:1px solid var(--border);padding:20px;border-radius:10px;margin-bottom:20px;}
            .chat-box{background:var(--card-bg);border:1px solid var(--border);border-radius:10px;height:400px;overflow-y:auto;padding:15px;margin-bottom:15px;}
            .message{margin:10px 0;padding:10px;border-radius:8px;max-width:80%;}
            .user{background:#1a1a2e;margin-right:auto;text-align:right;border:1px solid #FFD700;}
            .assistant{background:#1a1a2e;margin-left:auto;text-align:left;border:1px solid #4CAF50;}
            .input-area{display:flex;gap:10px;}
            .input-area input{flex:1;padding:10px;background:var(--bg);color:var(--text);border:1px solid var(--border);border-radius:5px;}
            .btn{background:var(--btn-bg);color:var(--btn-text);padding:10px 20px;border:none;border-radius:5px;cursor:pointer;}
            .theme-toggle{position:fixed;top:20px;left:20px;background:var(--btn-bg);color:var(--btn-text);border:none;border-radius:50%;width:50px;height:50px;font-size:20px;cursor:pointer;z-index:1000;}
            .model-badge{font-size:12px;padding:2px 8px;border-radius:10px;background:#4CAF50;color:#fff;margin-left:10px;}
            .typing{color:#aaa;font-style:italic;}
        </style>
    </head>
    <body>
        <button class="theme-toggle" onclick="toggleTheme()">🌓</button>
        <div class="container">
            <div class="header">
                <h2>🤖 المساعد الذكي المتقدم</h2>
                <p style="font-size:14px;color:#aaa;">مدعوم بـ ChatGPT مع RAG من قاعدة بيانات الأدوية</p>
                <div>
                    <span class="model-badge">ChatGPT</span>
                    <span class="model-badge" style="background:#FF9800;">RAG</span>
                </div>
            </div>
            <div class="chat-box" id="chatBox">
                <div class="message assistant">
                    👋 مرحباً! أنا مساعدك الصيدلاني الذكي. اسألني عن أي دواء، أو استفسار طبي، أو تحليل للبيانات.
                </div>
            </div>
            <div class="input-area">
                <input type="text" id="messageInput" placeholder="اكتب رسالتك هنا..." onkeypress="if(event.key==='Enter') sendMessage()">
                <button class="btn" onclick="sendMessage()">إرسال</button>
                <button class="btn" onclick="getInsights()" style="background:#4CAF50;color:#fff;">📊 تحليل</button>
            </div>
            <div style="margin-top:10px;display:flex;gap:10px;flex-wrap:wrap;">
                <button class="btn" onclick="quickQuestion('ما هو أفضل دواء مسكن؟')" style="font-size:12px;">💊 مسكنات</button>
                <button class="btn" onclick="quickQuestion('ما هي المضادات الحيوية المتوفرة؟')" style="font-size:12px;">🦠 مضادات حيوية</button>
                <button class="btn" onclick="quickQuestion('نصائح لتخزين الأدوية')" style="font-size:12px;">📦 تخزين</button>
                <button class="btn" onclick="quickQuestion('ما هي إيرادات الشهر الماضي؟')" style="font-size:12px;">💰 إيرادات</button>
            </div>
        </div>
        <script>
            function toggleTheme(){document.body.classList.toggle('light');localStorage.setItem('theme',document.body.classList.contains('light')?'light':'dark');}
            if(localStorage.getItem('theme')==='light') document.body.classList.add('light');

            function sendMessage() {
                const input = document.getElementById('messageInput');
                const msg = input.value.trim();
                if (!msg) return;
                
                addMessage('user', msg);
                input.value = '';
                addMessage('assistant', '⏳ جاري التفكير...', true);
                
                fetch('/api/chat', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({message: msg})
                })
                .then(r => r.json())
                .then(data => {
                    removeTyping();
                    if (data.success) {
                        let reply = data.reply;
                        if (data.model) {
                            reply += `\n\n🤖 (${data.model})`;
                        }
                        if (data.rag_results && data.rag_results.length > 0) {
                            reply += '\n\n📋 منتجات ذات صلة:';
                            data.rag_results.forEach((r, i) => {
                                if (i < 3) {
                                    reply += `\n• ${r.product.name} (تشابه: ${(r.similarity * 100).toFixed(0)}%)`;
                                }
                            });
                        }
                        addMessage('assistant', reply);
                    } else {
                        addMessage('assistant', '❌ عذراً، حدث خطأ: ' + (data.error || 'غير معروف'));
                    }
                })
                .catch(err => {
                    removeTyping();
                    addMessage('assistant', '❌ خطأ في الاتصال: ' + err);
                });
            }

            function getInsights() {
                addMessage('user', '📊 تحليل الأعمال');
                addMessage('assistant', '⏳ جاري تحليل البيانات...', true);
                
                fetch('/api/ai-insights', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({type: 'general'})
                })
                .then(r => r.json())
                .then(data => {
                    removeTyping();
                    if (data.success) {
                        addMessage('assistant', '📊 ' + data.insights);
                    } else {
                        addMessage('assistant', '❌ ' + (data.message || 'خطأ في التحليل'));
                    }
                })
                .catch(err => {
                    removeTyping();
                    addMessage('assistant', '❌ خطأ: ' + err);
                });
            }

            function quickQuestion(q) {
                document.getElementById('messageInput').value = q;
                sendMessage();
            }

            function addMessage(role, text, isTyping = false) {
                const chatBox = document.getElementById('chatBox');
                const div = document.createElement('div');
                div.className = `message ${role}`;
                if (isTyping) {
                    div.className += ' typing';
                    div.id = 'typingMessage';
                }
                div.textContent = text;
                chatBox.appendChild(div);
                chatBox.scrollTop = chatBox.scrollHeight;
            }

            function removeTyping() {
                const typing = document.getElementById('typingMessage');
                if (typing) typing.remove();
            }

            document.getElementById('messageInput').focus();
        </script>
    </body>
    </html>
    """)

# =============================== المسارات الإضافية (من الكود الأصلي) ===============================
@app.route('/admin/dynamic-pricing', methods=['GET', 'POST'])
@admin_required
def dynamic_pricing_page():
    if request.method == 'POST':
        product_id = request.form.get('product_id')
        if product_id:
            price = calculate_dynamic_price(int(product_id))
            return jsonify({'success': True, 'price': price})

    products = execute_query("""
        SELECT id, name, price, dynamic_price, sales_velocity, quantity, expiry_date,
               price_updated_at, abc_class
        FROM products WHERE is_active = 1 ORDER BY name
    """, fetch_all=True)

    update_sales_velocity()

    return render_template_string("""
    <!DOCTYPE html>
    <html dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>التسعير الديناميكي</title>
        <style>
            :root {
                --bg: #000;
                --text: #FFD700;
                --card-bg: #111;
                --border: #FFD700;
                --btn-bg: #FFD700;
                --btn-text: #000;
                --input-bg: #000;
                --input-text: #FFD700;
            }
            body.light {
                --bg: #f5f5f5;
                --text: #000;
                --card-bg: #fff;
                --border: #007bff;
                --btn-bg: #007bff;
                --btn-text: #fff;
                --input-bg: #fff;
                --input-text: #000;
            }
            body{background:var(--bg);color:var(--text);padding:20px;font-family:Arial;transition:background 0.3s,color 0.3s;}
            .container{max-width:1200px;margin:auto;}
            .header{background:var(--card-bg);border:1px solid var(--border);padding:20px;border-radius:10px;margin-bottom:20px;}
            .nav a{color:var(--text);text-decoration:none;margin-left:15px;}
            table{width:100%;border-collapse:collapse;background:var(--card-bg);border:1px solid var(--border);}
            th,td{padding:10px;border:1px solid var(--border);text-align:center;}
            th{background:var(--border);color:var(--btn-text);}
            .btn{background:var(--btn-bg);color:var(--btn-text);padding:8px 15px;border:none;border-radius:5px;cursor:pointer;}
            .price-up{color:#4CAF50;}
            .price-down{color:#f44336;}
            .theme-toggle{position:fixed;top:20px;left:20px;background:var(--btn-bg);color:var(--btn-text);border:none;border-radius:50%;width:50px;height:50px;font-size:20px;cursor:pointer;z-index:1000;}
            .chat-float{position:fixed;bottom:20px;left:20px;background:var(--btn-bg);color:var(--btn-text);width:60px;height:60px;border-radius:50%;font-size:30px;display:flex;align-items:center;justify-content:center;cursor:pointer;z-index:999;box-shadow:0 4px 15px rgba(0,0,0,0.5);text-decoration:none;}
        </style>
    </head>
    <body>
    <button class="theme-toggle" onclick="toggleTheme()">🌓</button>
    <a href="/chat" class="chat-float" title="المساعد الذكي">💬</a>
    <div class="container">
        <div class="header">
            <h2>📊 التسعير الديناميكي الذكي</h2>
            <div class="nav">
                <a href="/admin">لوحة المدير</a>
                <a href="/products">المنتجات</a>
                <a href="/admin/dynamic-pricing" style="font-weight:bold;">التسعير الديناميكي</a>
                <a href="/admin/ai-chat">🤖 المساعد الذكي</a>
            </div>
            <p style="margin-top:10px;font-size:14px;">يتم حساب السعر الأمثل تلقائياً بناءً على الطلب والمخزون وتاريخ الصلاحية والفئة</p>
            <button class="btn" onclick="updateAllPrices()">🔄 تحديث جميع الأسعار</button>
        </div>
        <div id="result"></div>
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>المنتج</th>
                    <th>السعر الحالي</th>
                    <th>السعر الديناميكي</th>
                    <th>سرعة المبيعات</th>
                    <th>المخزون</th>
                    <th>التصنيف</th>
                    <th>آخر تحديث</th>
                    <th>إجراء</th>
                </tr>
            </thead>
            <tbody>
                {% for p in products %}
                <tr>
                    <td>{{ p.id }}</td>
                    <td>{{ p.name }}</td>
                    <td>{{ p.price }} ريال</td>
                    <td class="{% if p.dynamic_price and p.dynamic_price > p.price %}price-up{% elif p.dynamic_price and p.dynamic_price < p.price %}price-down{% endif %}">
                        {{ p.dynamic_price or p.price }} ريال
                    </td>
                    <td>{{ "%.1f"|format(p.sales_velocity or 0) }}</td>
                    <td>{{ p.quantity }}</td>
                    <td>{{ p.abc_class or '-' }}</td>
                    <td>{{ p.price_updated_at or 'لم يحدث' }}</td>
                    <td>
                        <button class="btn" onclick="updatePrice({{ p.id }})">تحديث</button>
                    </td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    <script>
        function toggleTheme(){document.body.classList.toggle('light');localStorage.setItem('theme',document.body.classList.contains('light')?'light':'dark');}
        if(localStorage.getItem('theme')==='light') document.body.classList.add('light');

        function updatePrice(id) {
            fetch(`/api/dynamic-price/${id}`)
                .then(r => r.json())
                .then(data => {
                    if (data.success) {
                        document.getElementById('result').innerHTML = `<p style="color:green;">✅ تم تحديث السعر إلى ${data.price} ريال</p>`;
                        setTimeout(() => location.reload(), 1000);
                    }
                });
        }

        function updateAllPrices() {
            if (!confirm('تحديث جميع الأسعار قد يستغرق بعض الوقت. هل تريد المتابعة؟')) return;
            document.getElementById('result').innerHTML = '<p>⏳ جاري تحديث الأسعار...</p>';
            fetch('/api/update-all-prices', {method: 'POST'})
                .then(r => r.json())
                .then(data => {
                    document.getElementById('result').innerHTML = `<p style="color:green;">✅ ${data.message}</p>`;
                    setTimeout(() => location.reload(), 1500);
                });
        }
    </script>
    </body>
    </html>
    """, products=products)

@app.route('/api/update-all-prices', methods=['POST'])
@admin_required
def update_all_prices():
    products = execute_query("SELECT id FROM products WHERE is_active = 1", fetch_all=True)
    count = 0
    for p in products:
        calculate_dynamic_price(p['id'])
        count += 1
    return jsonify({'success': True, 'message': f'تم تحديث {count} منتج'})

def update_sales_velocity():
    products = execute_query("SELECT DISTINCT product_id FROM invoice_items", fetch_all=True)
    for p in products:
        pid = p['product_id']
        items = execute_query("""
            SELECT ii.quantity, i.created_at
            FROM invoice_items ii
            JOIN invoices i ON ii.invoice_id = i.id
            WHERE ii.product_id = ? AND i.created_at >= datetime('now', '-30 days')
        """, (pid,), fetch_all=True)
        if items:
            total_qty = sum(item['quantity'] for item in items)
            velocity = total_qty / 30.0
        else:
            velocity = 0.1
        execute_query("UPDATE products SET sales_velocity = ? WHERE id = ?", (velocity, pid), commit=True)

# =============================== باقي الصفحات الإدارية ===============================
@app.route('/admin/anomalies')
@admin_required
def anomalies_page():
    anomalies = execute_query("""
        SELECT i.*, u.username as created_by_name
        FROM invoices i
        LEFT JOIN users u ON i.created_by = u.id
        WHERE i.is_anomaly = 1
        ORDER BY i.created_at DESC
    """, fetch_all=True)

    anomaly_logs = execute_query("""
        SELECT al.*, i.invoice_number, i.customer_name
        FROM anomaly_logs al
        JOIN invoices i ON al.invoice_id = i.id
        ORDER BY al.created_at DESC LIMIT 50
    """, fetch_all=True)

    return render_template_string("""...""", anomalies=anomalies, anomaly_logs=anomaly_logs)

@app.route('/api/mark-anomaly-reviewed/<int:invoice_id>', methods=['POST'])
@admin_required
def mark_anomaly_reviewed(invoice_id):
    execute_query("UPDATE invoices SET is_anomaly = 0 WHERE id = ?", (invoice_id,), commit=True)
    execute_query("UPDATE anomaly_logs SET is_reviewed = 1, reviewed_by = ? WHERE invoice_id = ?",
                  (session.get('user_id', 1), invoice_id), commit=True)
    return jsonify({'success': True})

@app.route('/admin/scan-image-ui')
@login_required
def scan_image_ui():
    return render_template_string("""...""")

@app.route('/admin/scan-invoice')
@login_required
def scan_invoice_ui():
    suppliers = execute_query("SELECT id, name FROM suppliers", fetch_all=True)
    return render_template_string("""...""", suppliers=suppliers)

@app.route('/admin/feedback')
@admin_required
def feedback_page():
    feedbacks = execute_query("""
        SELECT * FROM feedback ORDER BY created_at DESC LIMIT 100
    """, fetch_all=True)

    stats = execute_query("""
        SELECT
            COUNT(*) as total,
            AVG(rating) as avg_rating,
            SUM(CASE WHEN sentiment_label = 'إيجابي' THEN 1 ELSE 0 END) as positive,
            SUM(CASE WHEN sentiment_label = 'سلبي' THEN 1 ELSE 0 END) as negative,
            SUM(CASE WHEN sentiment_label = 'محايد' THEN 1 ELSE 0 END) as neutral
        FROM feedback
    """, fetch_one=True)

    return render_template_string("""...""", stats=stats, feedbacks=feedbacks)

@app.route('/admin/analytics')
@admin_required
def analytics_page():
    stats = execute_query("""
        SELECT
            (SELECT COUNT(*) FROM products WHERE is_active = 1) as total_products,
            (SELECT COUNT(*) FROM customers) as total_customers,
            (SELECT COUNT(*) FROM invoices WHERE DATE(created_at) = DATE('now')) as today_sales,
            (SELECT IFNULL(SUM(final_total), 0) FROM invoices WHERE DATE(created_at) = DATE('now')) as today_revenue,
            (SELECT IFNULL(SUM(final_total), 0) FROM invoices WHERE created_at >= DATE('now', '-30 days')) as month_revenue
    """, fetch_one=True)

    top_products = execute_query("""
        SELECT p.name, SUM(ii.quantity) as total_sold, SUM(ii.total) as revenue
        FROM invoice_items ii
        JOIN products p ON ii.product_id = p.id
        GROUP BY ii.product_id
        ORDER BY total_sold DESC
        LIMIT 10
    """, fetch_all=True)

    category_sales = execute_query("""
        SELECT p.category, SUM(ii.total) as revenue
        FROM invoice_items ii
        JOIN products p ON ii.product_id = p.id
        WHERE p.category IS NOT NULL AND p.category != ''
        GROUP BY p.category
        ORDER BY revenue DESC
    """, fetch_all=True)

    daily_sales = execute_query("""
        SELECT DATE(created_at) as day, IFNULL(SUM(final_total), 0) as total
        FROM invoices
        WHERE created_at >= DATE('now', '-7 days')
        GROUP BY DATE(created_at)
        ORDER BY day
    """, fetch_all=True)

    return render_template_string("""...""", stats=stats, top_products=top_products, category_sales=category_sales, daily_sales=daily_sales)

# =============================== الصفحات الرئيسية (من الكود الأصلي) ===============================
@app.route('/')
def home():
    settings = get_settings()
    company_name = settings['company_name']
    company_logo = settings['company_logo']
    return render_template_string("""...""", company_name=company_name, company_logo=company_logo)

@app.route('/products')
def products_page():
    return render_template_string("""...""")

@app.route('/offers')
def offers_page():
    return render_template_string("""...""")

@app.route('/points')
def points_page():
    return render_template_string("""...""")

@app.route('/cart')
def cart_page():
    return render_template_string("""...""")

@app.route('/chat')
def chat_page():
    return render_template_string("""...""")

@app.route('/payment')
def payment_page():
    return render_template_string("""...""")

@app.route('/customer/recommendations')
def customer_recommendations_page():
    return render_template_string("""...""")

@app.route('/pos')
@login_required
def pos():
    return render_template_string("""...""")

@app.route('/pharmacist')
@pharmacist_required
def pharmacist_dashboard():
    return render_template_string("""...""")

@app.route('/stock')
@store_keeper_required
def stock_dashboard():
    return render_template_string("""...""")

# =============================== باقي المسارات الإدارية ===============================
@app.route('/admin')
@login_required
def admin_dashboard():
    if session['role'] != 'admin':
        return redirect(url_for('pos'))

    low_stock = execute_query("SELECT id, name, quantity, min_quantity FROM products WHERE quantity <= min_quantity AND is_active=1", fetch_all=True)
    low_stock_count = len(low_stock) if low_stock else 0

    anomalies_count = execute_query("SELECT COUNT(*) as count FROM invoices WHERE is_anomaly = 1", fetch_one=True)
    anomalies_count = anomalies_count['count'] if anomalies_count else 0

    feedback_count = execute_query("SELECT COUNT(*) as count FROM feedback WHERE DATE(created_at) = DATE('now')", fetch_one=True)
    feedback_count = feedback_count['count'] if feedback_count else 0

    return render_template_string("""
    <!DOCTYPE html><html dir="rtl"><head><title>لوحة الإدارة</title>
    <style>
        :root {
            --bg: #000;
            --text: #FFD700;
            --card-bg: #111;
            --border: #FFD700;
            --btn-bg: #FFD700;
            --btn-text: #000;
        }
        body.light {
            --bg: #f5f5f5;
            --text: #000;
            --card-bg: #fff;
            --border: #007bff;
            --btn-bg: #007bff;
            --btn-text: #fff;
        }
        body{background:var(--bg);color:var(--text);font-family:Arial;padding:20px;transition:background 0.3s,color 0.3s;}
        .header{background:var(--card-bg);border:1px solid var(--border);padding:20px;border-radius:10px;margin-bottom:20px;}
        .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:15px;}
        .card{background:var(--card-bg);border:1px solid var(--border);padding:18px;border-radius:10px;text-align:center;cursor:pointer;transition:0.3s;color:var(--text);text-decoration:none;display:block;}
        .card:hover{transform:translateY(-5px);background:#222;}
        .card .icon{font-size:36px;display:block;margin-bottom:10px;}
        .card .title{font-weight:bold;}
        .card .badge{background:#f44336;color:#fff;border-radius:50%;padding:2px 10px;font-size:12px;margin-right:5px;}
        .logout{position:absolute;top:20px;left:20px;background:var(--btn-bg);color:var(--btn-text);padding:10px;border-radius:5px;text-decoration:none;}
        .alert{background:#8B0000;border:1px solid var(--border);padding:10px;border-radius:5px;margin-bottom:20px;color:#fff;}
        .chat-float{position:fixed;bottom:20px;left:20px;background:var(--btn-bg);color:var(--btn-text);width:60px;height:60px;border-radius:50%;font-size:30px;display:flex;align-items:center;justify-content:center;cursor:pointer;z-index:999;box-shadow:0 4px 15px rgba(0,0,0,0.5);text-decoration:none;}
        .theme-toggle{position:fixed;top:20px;left:70px;background:var(--btn-bg);color:var(--btn-text);border:none;border-radius:50%;width:50px;height:50px;font-size:20px;cursor:pointer;z-index:1000;}
        .section-title{margin:25px 0 15px 0;border-right:4px solid var(--border);padding-right:15px;}
        .ai-badge{background:#4CAF50;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;}
    </style>
    </head>
    <body>
    <button class="theme-toggle" onclick="toggleTheme()">🌓</button>
    <a href="/chat" class="chat-float" title="المساعد الذكي">💬</a>
    <a href="/logout" class="logout">تسجيل خروج</a>
    <div class="header">
        <h1>🎛️ لوحة التحكم - وكالة البشائر</h1>
        <p>مرحباً {{ session.username }} (مدير) <span class="ai-badge">🧠 AI</span></p>
        <p style="font-size:14px;color:#aaa;">نظام متكامل مع ميزات الذكاء الاصطناعي - ChatGPT + RAG</p>
    </div>

    {% if low_stock_count > 0 %}
    <div class="alert">⚠️ تنبيه: يوجد {{ low_stock_count }} منتج (منتجات) مخزونها منخفض!</div>
    {% endif %}
    {% if anomalies_count > 0 %}
    <div class="alert" style="background:#8B0000;">🚨 تنبيه: يوجد {{ anomalies_count }} فاتورة شاذة تحتاج إلى مراجعة!</div>
    {% endif %}
    {% if feedback_count > 0 %}
    <div class="alert" style="background:#004d40;">💬 يوجد {{ feedback_count }} تقييم جديد اليوم!</div>
    {% endif %}

    <h3 class="section-title">📋 الإدارة الأساسية</h3>
    <div class="grid">
        <a href="/admin/settings" class="card"><span class="icon">⚙️</span><span class="title">الإعدادات</span></a>
        <a href="/admin/products" class="card"><span class="icon">📦</span><span class="title">المنتجات</span></a>
        <a href="/admin/suppliers" class="card"><span class="icon">🏭</span><span class="title">الموردين</span></a>
        <a href="/admin/purchases" class="card"><span class="icon">📥</span><span class="title">المشتريات</span></a>
        <a href="/admin/offers" class="card"><span class="icon">🎁</span><span class="title">العروض</span></a>
        <a href="/admin/customers" class="card"><span class="icon">👥</span><span class="title">العملاء</span></a>
        <a href="/admin/invoices" class="card"><span class="icon">📄</span><span class="title">الفواتير</span></a>
        <a href="/admin/reports" class="card"><span class="icon">📊</span><span class="title">التقارير</span></a>
        <a href="/admin/users" class="card"><span class="icon">👤</span><span class="title">المستخدمين</span></a>
        <a href="/admin/returns" class="card"><span class="icon">🔄</span><span class="title">المرتجعات</span></a>
    </div>

    <h3 class="section-title">🧠 ميزات الذكاء الاصطناعي</h3>
    <div class="grid">
        <a href="/admin/ai-chat" class="card" style="border-color:#4CAF50;"><span class="icon">🤖</span><span class="title">المساعد الذكي ChatGPT</span></a>
        <a href="/admin/dynamic-pricing" class="card"><span class="icon">💰</span><span class="title">التسعير الديناميكي</span></a>
        <a href="/admin/anomalies" class="card"><span class="icon">🛡️</span><span class="title">كشف الشذوذ <span class="badge">{{ anomalies_count }}</span></span></a>
        <a href="/admin/scan-image-ui" class="card"><span class="icon">📷</span><span class="title">التعرف من الصور</span></a>
        <a href="/admin/scan-invoice" class="card"><span class="icon">📄</span><span class="title">استخراج الفواتير</span></a>
        <a href="/admin/feedback" class="card"><span class="icon">💬</span><span class="title">تحليل المشاعر <span class="badge">{{ feedback_count }}</span></span></a>
        <a href="/admin/analytics" class="card"><span class="icon">📈</span><span class="title">التحليل الذكي</span></a>
        <a href="/customer/recommendations" class="card"><span class="icon">🎯</span><span class="title">توصيات العملاء</span></a>
    </div>

    <h3 class="section-title">🔗 روابط سريعة</h3>
    <div class="grid">
        <a href="/pos" class="card"><span class="icon">🛒</span><span class="title">نقطة البيع</span></a>
        <a href="/payment" class="card"><span class="icon">💳</span><span class="title">السداد الإلكتروني</span></a>
        <a href="/admin/reports/export/inventory?format=excel" class="card"><span class="icon">📊</span><span class="title">تصدير المخزون Excel</span></a>
    </div>

    <script>
        function toggleTheme(){document.body.classList.toggle('light');localStorage.setItem('theme',document.body.classList.contains('light')?'light':'dark');}
        if(localStorage.getItem('theme')==='light') document.body.classList.add('light');
    </script>
    </body>
    </html>
    """, low_stock_count=low_stock_count, anomalies_count=anomalies_count, feedback_count=feedback_count)

# =============================== تسجيل الدخول والخروج ===============================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = execute_query("SELECT id, username, password_hash, role FROM users WHERE username = ?", (username,), fetch_one=True)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user['role'] == 'cashier':
                return redirect(url_for('pos'))
            elif user['role'] == 'pharmacist':
                return redirect(url_for('pharmacist_dashboard'))
            else:
                return redirect(url_for('stock_dashboard'))
        else:
            return render_template_string("<h2 style='color:red;text-align:center;'>بيانات دخول خاطئة</h2><a href='/login'>حاول مرة أخرى</a>")
    return render_template_string("""
    <!DOCTYPE html><html dir="rtl"><head><title>تسجيل الدخول</title>
    <style>body{background:#000;color:#FFD700;font-family:Arial;padding:50px;}.login{max-width:400px;margin:auto;background:#111;padding:30px;border-radius:10px;border:1px solid #FFD700;}</style>
    </head><body><div class="login"><h2>تسجيل الدخول</h2><form method="post"><input type="text" name="username" placeholder="اسم المستخدم" required style="width:100%;padding:10px;margin:10px 0;background:#000;color:#FFD700;border:1px solid #FFD700;"><input type="password" name="password" placeholder="كلمة المرور" required style="width:100%;padding:10px;margin:10px 0;background:#000;color:#FFD700;border:1px solid #FFD700;"><button type="submit" style="background:#FFD700;color:#000;padding:10px;width:100%;border:none;">دخول</button></form></div></body>
    """)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

# =============================== API إضافية ===============================
@app.route('/api/products')
def api_products():
    limit = request.args.get('limit', 50, type=int)
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    query = """
        SELECT p.*, u.name as unit_name, u.symbol as unit_symbol
        FROM products p
        LEFT JOIN units u ON p.unit_id = u.id
        WHERE p.is_active = 1
    """
    params = []
    if search:
        query += " AND (p.name LIKE ? OR p.description LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])
    if category:
        query += " AND p.category = ?"
        params.append(category)
    query += " ORDER BY p.name LIMIT ?"
    params.append(limit)
    products = execute_query(query, params, fetch_all=True)
    return jsonify({'success': True, 'products': [dict(p) for p in products]})

@app.route('/api/categories')
def api_categories():
    rows = execute_query("SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != '' AND is_active = 1", fetch_all=True)
    categories = [row['category'] for row in rows] if rows else []
    return jsonify({'categories': categories})

@app.route('/api/offers')
def api_offers():
    offers = execute_query("SELECT * FROM offers WHERE is_active = 1 AND (start_date <= date('now') OR start_date IS NULL) AND (end_date >= date('now') OR end_date IS NULL)", fetch_all=True)
    return jsonify({'offers': [dict(o) for o in offers]})

@app.route('/api/customer/<phone>')
def api_customer(phone):
    customer = execute_query("SELECT * FROM customers WHERE phone = ?", (phone,), fetch_one=True)
    if customer:
        return jsonify({'success': True, **dict(customer)})
    return jsonify({'success': False, 'message': 'العميل غير موجود'})

@app.route('/api/scan-barcode', methods=['POST'])
def api_scan_barcode():
    data = request.json
    barcode = data.get('barcode', '').strip()
    if not barcode:
        return jsonify({'success': False, 'message': 'الباركود فارغ'})
    product = execute_query("""
        SELECT p.*, u.name as unit_name
        FROM products p
        LEFT JOIN units u ON p.unit_id = u.id
        WHERE p.barcode = ? AND p.is_active = 1
    """, (barcode,), fetch_one=True)
    if product:
        return jsonify({'success': True, 'product': dict(product)})
    return jsonify({'success': False, 'message': 'المنتج غير موجود'})

@app.route('/api/create_invoice', methods=['POST'])
def create_invoice():
    data = request.json
    try:
        cart = data.get('cart', [])
        customer_name = data.get('customer_name', '')
        customer_phone = data.get('customer_phone', '')
        customer_address = data.get('customer_address', '')
        payment_method = data.get('payment_method', 'cash')
        total = sum(item['price'] * item['quantity'] for item in cart)
        discount = data.get('discount', 0)
        final_total = total - discount
        invoice_number = f"INV-{int(time.time())}"

        invoice_data = {'total': total, 'discount': discount, 'final_total': final_total, 'item_count': len(cart)}
        anomaly_result = detect_anomaly(invoice_data)

        execute_query("""
            INSERT INTO invoices (invoice_number, customer_name, customer_phone, customer_address, total, discount, final_total, payment_method, created_by, is_anomaly, anomaly_score, anomaly_reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (invoice_number, customer_name, customer_phone, customer_address, total, discount, final_total, payment_method, session.get('user_id', 1),
              anomaly_result['is_anomaly'], anomaly_result['score'], anomaly_result['reason']), commit=True)

        invoice = execute_query("SELECT id FROM invoices WHERE invoice_number = ?", (invoice_number,), fetch_one=True)
        invoice_id = invoice['id']

        if anomaly_result['is_anomaly']:
            execute_query("""
                INSERT INTO anomaly_logs (invoice_id, anomaly_score, reason)
                VALUES (?, ?, ?)
            """, (invoice_id, anomaly_result['score'], anomaly_result['reason']), commit=True)

        for item in cart:
            execute_query("""
                INSERT INTO invoice_items (invoice_id, product_id, product_name, quantity, price, total)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (invoice_id, item['id'], item['name'], item['quantity'], item['price'], item['price'] * item['quantity']), commit=True)

            product = execute_query("SELECT quantity FROM products WHERE id = ?", (item['id'],), fetch_one=True)
            if product:
                new_qty = product['quantity'] - item['quantity']
                execute_query("UPDATE products SET quantity = ? WHERE id = ?", (new_qty, item['id']), commit=True)
                log_inventory(item['id'], item['name'], 'sale', -item['quantity'], product['quantity'], new_qty, f"فاتورة {invoice_number}", session.get('user_id', 1))

        return jsonify({'success': True, 'invoice_number': invoice_number, 'is_anomaly': anomaly_result['is_anomaly']})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# =============================== تصدير التقارير ===============================
@app.route('/admin/reports/export/<report_type>')
@login_required
def export_report(report_type):
    if report_type == 'inventory':
        data = execute_query("""
            SELECT p.id, p.name, p.category, p.quantity, p.min_quantity, p.price, p.cost_price,
                   p.expiry_date, p.batch_number, p.manufacturer, u.name as unit,
                   p.dynamic_price, p.sales_velocity, p.abc_class
            FROM products p
            LEFT JOIN units u ON p.unit_id = u.id
            WHERE p.is_active = 1
            ORDER BY p.name
        """, fetch_all=True)
        title = 'تقرير المخزون'
        headers = ['الرقم', 'الاسم', 'الفئة', 'الكمية', 'الحد الأدنى', 'سعر البيع', 'سعر الشراء', 'تاريخ الصلاحية', 'رقم الدفعة', 'الشركة المصنعة', 'الوحدة', 'السعر الديناميكي', 'سرعة المبيعات', 'تصنيف ABC']
    elif report_type == 'sales':
        data = execute_query("""
            SELECT i.invoice_number, i.customer_name, i.final_total, i.payment_method, i.created_at,
                   (SELECT COUNT(*) FROM invoice_items WHERE invoice_id = i.id) as items_count,
                   i.is_anomaly
            FROM invoices i
            ORDER BY i.created_at DESC
            LIMIT 100
        """, fetch_all=True)
        title = 'تقرير المبيعات (آخر 100 فاتورة)'
        headers = ['رقم الفاتورة', 'العميل', 'الإجمالي', 'طريقة الدفع', 'التاريخ', 'عدد الأصناف', 'شاذ']
    elif report_type == 'expiry':
        data = execute_query("""
            SELECT id, name, expiry_date, quantity, batch_number, manufacturer
            FROM products
            WHERE expiry_date IS NOT NULL AND is_active = 1
            ORDER BY expiry_date ASC
        """, fetch_all=True)
        title = 'تقرير الصلاحية (الأقرب للانتهاء أولاً)'
        headers = ['الرقم', 'الاسم', 'تاريخ الصلاحية', 'الكمية', 'رقم الدفعة', 'الشركة المصنعة']
    else:
        return 'تقرير غير معروف', 400

    if not data:
        return 'لا توجد بيانات', 404

    fmt = request.args.get('format', 'excel')
    if fmt == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = title
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                if isinstance(value, (datetime.date, datetime.datetime)):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name=f"{title}.xlsx", as_attachment=True)
    elif fmt == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.alignment = 1
        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.5*cm))
        table_data = [headers]
        for row in data:
            row_list = []
            for value in row:
                if isinstance(value, (datetime.date, datetime.datetime)):
                    row_list.append(value.isoformat())
                else:
                    row_list.append(str(value) if value is not None else '')
            table_data.append(row_list)
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, download_name=f"{title}.pdf", as_attachment=True)
    else:
        return 'صيغة غير مدعومة', 400

# =============================== الصفحات الإدارية (قيد التطوير) ===============================
@app.route('/admin/settings')
@admin_required
def admin_settings():
    return render_template_string("""
    <!DOCTYPE html><html dir="rtl"><head><title>الإعدادات</title>
    <style>body{background:#000;color:#FFD700;padding:20px;font-family:Arial;}</style>
    </head><body>
    <h2>⚙️ الإعدادات</h2>
    <p>قيد التطوير...</p>
    <a href="/admin">العودة للوحة</a>
    </body></html>
    """)

@app.route('/admin/products')
@admin_required
def admin_products():
    return render_template_string("""
    <!DOCTYPE html><html dir="rtl"><head><title>المنتجات</title>
    <style>body{background:#000;color:#FFD700;padding:20px;font-family:Arial;}</style>
    </head><body>
    <h2>📦 إدارة المنتجات</h2>
    <p>قيد التطوير...</p>
    <a href="/admin">العودة للوحة</a>
    </body></html>
    """)

@app.route('/admin/suppliers')
@admin_required
def admin_suppliers():
    return render_template_string("""...""")

@app.route('/admin/purchases')
@admin_required
def admin_purchases():
    return render_template_string("""...""")

@app.route('/admin/offers')
@admin_required
def admin_offers():
    return render_template_string("""...""")

@app.route('/admin/customers')
@admin_required
def admin_customers():
    return render_template_string("""...""")

@app.route('/admin/invoices')
@admin_required
def admin_invoices():
    return render_template_string("""...""")

@app.route('/admin/reports')
@admin_required
def admin_reports():
    return render_template_string("""...""")

@app.route('/admin/users')
@admin_required
def admin_users():
    return render_template_string("""...""")

@app.route('/admin/returns')
@admin_required
def admin_returns():
    return render_template_string("""...""")

# =============================== تشغيل التطبيق ===============================
if __name__ == '__main__':
    print("="*70)
    print("🚀 وكالة البشائر للأدوية والمستلزمات الطبية (نسخة متطورة بالذكاء الاصطناعي)")
    print("="*70)
    print(f"📁 قاعدة البيانات: {'PostgreSQL' if DATABASE_URL else 'SQLite local'}")
    print("🧠 ميزات الذكاء الاصطناعي:")
    print("   🤖 مساعد ذكي متقدم (ChatGPT + RAG)")
    print("   💰 التسعير الديناميكي (Dynamic Pricing)")
    print("   🛡️ كشف الشذوذ والاحتيال (Anomaly Detection)")
    print("   👁️ تحليل الصور (GPT-4 Vision)")
    print("   📷 التعرف على الأدوية من الصور (OCR + CV)")
    print("   🎯 توصيات مخصصة للعملاء")
    print("   💬 تحليل المشاعر من التقييمات")
    print("   📄 استخراج بيانات الفواتير الورقية")
    print("   📊 لوحات تحليل ذكية")
    print("="*70)
    print("🔐 بيانات الدخول:")
    print("   admin / admin123 (مدير)")
    print("   pharmacist / pharma123 (صيدلي)")
    print("   cashier / cashier123 (كاشير)")
    print("   stock / stock123 (أمين مخزن)")
    print("   purchaser / purch123 (مندوب مشتريات)")
    print("="*70)
    print("🌐 الروابط الرئيسية:")
    print("   👉 http://localhost:5000/                (الرئيسية)")
    print("   👉 http://localhost:5000/admin           (لوحة المدير - جميع الميزات)")
    print("   👉 http://localhost:5000/admin/ai-chat   (المساعد الذكي ChatGPT)")
    print("   👉 http://localhost:5000/admin/dynamic-pricing (التسعير الديناميكي)")
    print("   👉 http://localhost:5000/admin/anomalies (كشف الشذوذ)")
    print("   👉 http://localhost:5000/admin/scan-image-ui (التعرف من الصور)")
    print("   👉 http://localhost:5000/admin/scan-invoice (استخراج الفواتير)")
    print("   👉 http://localhost:5000/admin/feedback  (تحليل المشاعر)")
    print("   👉 http://localhost:5000/admin/analytics (التحليل الذكي)")
    print("="*70)
    print("✅ تم التحميل بنجاح! افتح الرابط في المتصفح.")
    app.run(host='127.0.0.1', port=5000, debug=True, threaded=True)
